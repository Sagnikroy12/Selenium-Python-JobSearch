import math
import os
import re
import smtplib
import ssl
from collections import Counter
from dataclasses import dataclass
from email.message import EmailMessage
from html import escape
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader
from config.config import config
from sentence_transformers import SentenceTransformer, util
from sklearn.feature_extraction.text import ENGLISH_STOP_WORDS


EMBEDDING_MODEL = SentenceTransformer("all-MiniLM-L6-v2")


ATS_SCORE_COLUMN = "ATS Score"
SEMANTIC_MATCH_COLUMN = "Semantic Match"
SKILL_MATCH_COLUMN = "Skill Match"
KEYWORD_MATCH_COLUMN = "Keyword Match"
TITLE_MATCH_COLUMN = "Title Match"
EXPERIENCE_MATCH_COLUMN = "Experience Match"
MATCHED_SKILLS_COLUMN = "Matched Skills"
MISSING_SKILLS_COLUMN = "Missing Skills"
RECOMMENDATION_COLUMN = "Recommendation"
JOB_LINK_COLUMN = "Job Link"
JOB_DESCRIPTION_COLUMN = "Job Description"
JOB_TITLE_COLUMN = "Job Title"

ATS_COLUMNS = [
    ATS_SCORE_COLUMN,
    SEMANTIC_MATCH_COLUMN,
    SKILL_MATCH_COLUMN,
    KEYWORD_MATCH_COLUMN,
    TITLE_MATCH_COLUMN,
    EXPERIENCE_MATCH_COLUMN,
    MATCHED_SKILLS_COLUMN,
    MISSING_SKILLS_COLUMN,
    RECOMMENDATION_COLUMN,
]

TECH_SKILLS = {
    "accessibility",
    "agile",
    "airflow",
    "api",
    "appium",
    "aws",
    "azure",
    "bdd",
    "browserstack",
    "ci/cd",
    "cypress",
    "docker",
    "etl",
    "excel",
    "fastapi",
    "flask",
    "git",
    "github actions",
    "graphql",
    "java",
    "javascript",
    "jenkins",
    "jira",
    "jmeter",
    "kubernetes",
    "linux",
    "manual testing",
    "microservices",
    "mongodb",
    "mysql",
    "playwright",
    "postgresql",
    "postman",
    "pytest",
    "python",
    "qa",
    "rest api",
    "robot framework",
    "selenium",
    "sql",
    "test automation",
    "testng",
    "typescript",
    "unix",
}

CERTIFICATION_TERMS = {
    "aws certified",
    "azure certified",
    "bachelor",
    "b.tech",
    "btech",
    "certified",
    "computer science",
    "degree",
    "istqb",
    "master",
}

WORD_PATTERN = re.compile(r"[a-z][a-z0-9+#./-]*")
YEARS_PATTERN = re.compile(
    r"(?P<years>\d+(?:\.\d+)?)\s*\+?\s*(?:years?|yrs?)\s*(?:of)?\s*(?:experience|exp)?",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class ATSScore:
    ats_score: float
    semantic_match: float
    skill_match: float
    keyword_match: float
    title_match: float
    experience_match: float
    education_bonus: float
    matched_skills: list[str]
    missing_skills: list[str]
    recommendation: str

    def as_columns(self):
        return {
            ATS_SCORE_COLUMN: self.ats_score,
            SEMANTIC_MATCH_COLUMN: self.semantic_match,
            SKILL_MATCH_COLUMN: self.skill_match,
            KEYWORD_MATCH_COLUMN: self.keyword_match,
            TITLE_MATCH_COLUMN: self.title_match,
            EXPERIENCE_MATCH_COLUMN: self.experience_match,
            MATCHED_SKILLS_COLUMN: ", ".join(self.matched_skills),
            MISSING_SKILLS_COLUMN: ", ".join(self.missing_skills),
            RECOMMENDATION_COLUMN: self.recommendation,
        }


def extract_resume_text(resume_pdf_path):
    path = Path(resume_pdf_path)
    if not path.exists():
        raise FileNotFoundError(f"Resume PDF not found: {path}")

    reader = PdfReader(str(path))
    text_parts = []
    for page in reader.pages:
        text_parts.append(page.extract_text() or "")

    resume_text = "\n".join(text_parts).strip()
    if not resume_text:
        raise ValueError(f"No readable text could be extracted from resume PDF: {path}")
    return resume_text


def normalize_text(value):
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).lower()).strip()


def calculate_semantic_match(resume_text, job_description, resume_embedding=None):
    resume = normalize_text(resume_text)
    description = normalize_text(job_description)
    if not resume or not description:
        return 0.0

    try:
        if resume_embedding is None:
            resume_embedding = EMBEDDING_MODEL.encode(resume, convert_to_tensor=True)
        job_embedding = EMBEDDING_MODEL.encode(description, convert_to_tensor=True)
        similarity = float(util.cos_sim(resume_embedding, job_embedding).item())
        return bounded_percentage(similarity * 100)
    except (ValueError, RuntimeError):
        return 0.0


def calculate_match_percentage(resume_text, job_description):
    return calculate_semantic_match(resume_text, job_description)


def calculate_ats_score(resume_text, job_description, job_title="", resume_embedding=None):
    resume = normalize_text(resume_text)
    description = normalize_text(job_description)
    title = normalize_text(job_title)

    if not description:
        return ATSScore(
            ats_score=0.0,
            semantic_match=0.0,
            skill_match=0.0,
            keyword_match=0.0,
            title_match=0.0,
            experience_match=0.0,
            education_bonus=0.0,
            matched_skills=[],
            missing_skills=[],
            recommendation="No job description available for scoring.",
        )

    semantic_match = calculate_semantic_match(
        resume,
        description,
        resume_embedding=resume_embedding,
    )
    job_skills = extract_skills(description)
    resume_skills = extract_skills(resume)
    matched_skills = sorted(job_skills & resume_skills)
    missing_skills = sorted(job_skills - resume_skills)
    skill_match = percentage(len(matched_skills), len(job_skills))
    keyword_match = calculate_keyword_match(resume, description)
    title_match = calculate_title_match(resume, title)
    experience_match = calculate_experience_match(resume, description)
    education_bonus = calculate_education_bonus(resume, description)

    weighted_score = (
        semantic_match * 0.25
        + skill_match * 0.30
        + keyword_match * 0.20
        + title_match * 0.10
        + experience_match * 0.10
        + education_bonus * 0.05
    )


    ats_score = bounded_percentage(weighted_score)
    return ATSScore(
        ats_score=ats_score,
        semantic_match=semantic_match,
        skill_match=skill_match,
        keyword_match=keyword_match,
        title_match=title_match,
        experience_match=experience_match,
        education_bonus=education_bonus,
        matched_skills=matched_skills,
        missing_skills=missing_skills[:12],
        recommendation=build_recommendation(ats_score, missing_skills),
    )


def extract_skills(text):
    normalized = normalize_text(text)
    return {skill for skill in TECH_SKILLS if skill in normalized}


def calculate_keyword_match(resume_text, job_description, top_n=25):
    keywords = extract_weighted_keywords(job_description, top_n=top_n)
    if not keywords:
        return 0.0

    matched_weight = sum(weight for keyword, weight in keywords if keyword in resume_text)
    total_weight = sum(weight for _, weight in keywords)
    return percentage(matched_weight, total_weight)


def extract_weighted_keywords(text, top_n=25):
    tokens = [
        token
        for token in WORD_PATTERN.findall(normalize_text(text))
        if len(token) > 2 and token not in ENGLISH_STOP_WORDS
    ]
    if not tokens:
        return []

    counts = Counter(tokens)
    for skill in extract_skills(text):
        counts[skill] += 2
    return counts.most_common(top_n)


def calculate_title_match(resume_text, job_title):
    title_tokens = {
        token
        for token in WORD_PATTERN.findall(job_title)
        if len(token) > 2 and token not in ENGLISH_STOP_WORDS
    }
    if not title_tokens:
        return 0.0
    matched_tokens = {token for token in title_tokens if token in resume_text}
    return percentage(len(matched_tokens), len(title_tokens))


def calculate_experience_match(resume_text, job_description):
    required_years = extract_max_years(job_description)
    if required_years is None:
        return 100.0

    if config.user_experience_years > 0:
        resume_years = config.user_experience_years
    else:
        resume_years = extract_max_years(resume_text)

    if resume_years is None:
        return 0.0

    if resume_years >= required_years:
        return 100.0

    return bounded_percentage((resume_years / required_years) * 100)


def extract_max_years(text):
    years = [float(match.group("years")) for match in YEARS_PATTERN.finditer(text)]
    if not years:
        return None
    return max(years)


def calculate_education_bonus(resume_text, job_description):
    required_terms = {term for term in CERTIFICATION_TERMS if term in job_description}
    if not required_terms:
        return 100.0

    matched_terms = {term for term in required_terms if term in resume_text}
    return percentage(len(matched_terms), len(required_terms))


def build_recommendation(ats_score, missing_skills):
    if ats_score >= 80:
        return "Strong match. Apply soon and tailor the opening summary."
    if ats_score >= 65:
        return "Good match. Add missing priority skills before applying."
    if ats_score >= 45:
        return "Moderate match. Customize resume bullets for this role."
    if missing_skills:
        return "Low match. Focus on missing skills before applying."
    return "Low match. Review the job description before applying."


def percentage(numerator, denominator):
    if not denominator:
        return 0.0
    return bounded_percentage((numerator / denominator) * 100)


def bounded_percentage(value):
    if value is None or math.isnan(float(value)):
        return 0.0
    return round(max(0.0, min(100.0, float(value))), 2)


def score_jobs_excel(excel_path, resume_pdf_path):
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Scraped jobs Excel file not found: {path}")

    resume_text = extract_resume_text(resume_pdf_path)
    resume_embedding = EMBEDDING_MODEL.encode(normalize_text(resume_text), convert_to_tensor=True)
    df = pd.read_excel(path)

    if JOB_DESCRIPTION_COLUMN not in df.columns:
        raise ValueError("Expected column 'Job Description' was not found in scraped jobs workbook.")

    score_rows = []
    for _, row in df.iterrows():
        score = calculate_ats_score(
            resume_text=resume_text,
            job_description=row.get(JOB_DESCRIPTION_COLUMN, ""),
            job_title=row.get(JOB_TITLE_COLUMN, ""),
            resume_embedding=resume_embedding,
        )
        score_rows.append(score.as_columns())

    score_df = pd.DataFrame(score_rows)
    for column in ATS_COLUMNS:
        df[column] = score_df[column]

    scored_jobs = df.sort_values(ATS_SCORE_COLUMN, ascending=False)
    top_matches = scored_jobs.head(10)
    scored_jobs.to_excel(path, index=False)
    apply_job_link_hyperlinks(path)
    return top_matches


def apply_job_link_hyperlinks(excel_path):
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    if JOB_LINK_COLUMN not in headers:
        workbook.save(excel_path)
        return

    link_column = headers.index(JOB_LINK_COLUMN) + 1
    for row in worksheet.iter_rows(min_row=2, min_col=link_column, max_col=link_column):
        cell = row[0]
        if not cell.value:
            continue
        cell.hyperlink = str(cell.value)
        cell.style = "Hyperlink"

    workbook.save(excel_path)


def build_top_matches_html(top_matches):
    rows = []
    for _, row in top_matches.iterrows():
        job_title = escape(str(row.get(JOB_TITLE_COLUMN, "")))
        job_link = row.get(JOB_LINK_COLUMN, "")
        if pd.notna(job_link) and str(job_link).strip():
            title_cell = (
                f'<a href="{escape(str(job_link).strip(), quote=True)}" '
                'style="color:#0a66c2;text-decoration:underline;">'
                f"{job_title}</a>"
            )
        else:
            title_cell = job_title

        rows.append(
            "<tr>"
            f"<td style=\"padding:10px;border-bottom:1px solid #dfe5ee;\">{title_cell}</td>"
            f"<td style=\"padding:10px;border-bottom:1px solid #dfe5ee;\">"
            f"{escape(str(row.get(ATS_SCORE_COLUMN, 0.0)))}%</td>"
            f"<td style=\"padding:10px;border-bottom:1px solid #dfe5ee;\">"
            f"{escape(str(row.get(MATCHED_SKILLS_COLUMN, '')))}</td>"
            f"<td style=\"padding:10px;border-bottom:1px solid #dfe5ee;\">"
            f"{escape(str(row.get(RECOMMENDATION_COLUMN, '')))}</td>"
            "</tr>"
        )

    table_rows = "".join(rows) or (
        '<tr><td colspan="4">No jobs were available for scoring.</td></tr>'
    )
    return f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f5f7fb;font-family:Arial,sans-serif;color:#172033;">
    <div style="max-width:980px;margin:0 auto;padding:28px 16px;">
      <h2 style="margin:0 0 8px;font-size:22px;">Daily ATS Job Match Report</h2>
      <p style="margin:0 0 20px;color:#516070;">Top 10 roles ranked by deterministic ATS-style scoring.</p>
      <table style="width:100%;border-collapse:collapse;background:#ffffff;border:1px solid #dfe5ee;">
        <thead>
          <tr>
            <th style="text-align:left;padding:12px;border-bottom:1px solid #dfe5ee;background:#eef3f8;">Job Title</th>
            <th style="text-align:left;padding:12px;border-bottom:1px solid #dfe5ee;background:#eef3f8;">ATS Score</th>
            <th style="text-align:left;padding:12px;border-bottom:1px solid #dfe5ee;background:#eef3f8;">Matched Skills</th>
            <th style="text-align:left;padding:12px;border-bottom:1px solid #dfe5ee;background:#eef3f8;">Recommendation</th>
          </tr>
        </thead>
        <tbody>{table_rows}</tbody>
      </table>
    </div>
  </body>
</html>
"""


def required_env(name):
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def send_top_matches_email(top_matches, recipient_email):
    smtp_host = required_env("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_username = required_env("SMTP_USERNAME")
    smtp_password = required_env("SMTP_PASSWORD")
    sender_email = os.getenv("MAIL_FROM") or smtp_username

    message = EmailMessage()
    message["Subject"] = os.getenv("EMAIL_SUBJECT", "Daily ATS Job Match Report")
    message["From"] = sender_email
    message["To"] = recipient_email
    message.set_content("Your daily ATS job match report is ready. Please view this email in HTML.")
    message.add_alternative(build_top_matches_html(top_matches), subtype="html")

    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_host, smtp_port) as smtp:
        smtp.starttls(context=context)
        smtp.login(smtp_username, smtp_password)
        smtp.send_message(message)


def run_ats_pipeline(excel_path, resume_pdf_path, recipient_email, send_email=True):
    top_matches = score_jobs_excel(excel_path, resume_pdf_path)
    if send_email:
        send_top_matches_email(top_matches, recipient_email)
    return top_matches
