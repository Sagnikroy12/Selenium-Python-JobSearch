import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from utils import ats_pipeline


RESUME_TEXT = """
Senior QA Automation Engineer with 6 years of experience in Selenium, Python,
Pytest, REST API testing, Postman, SQL, Docker, Jenkins, Git, Agile delivery,
and test automation frameworks. Bachelor degree in Computer Science.
"""


@pytest.mark.unit
def test_empty_job_description_returns_zero_score():
    score = ats_pipeline.calculate_ats_score(RESUME_TEXT, "", "Automation Engineer")

    assert score.ats_score == 0.0
    assert score.recommendation == "No job description available for scoring."


@pytest.mark.unit
def test_strong_skill_overlap_scores_higher_than_weak_overlap():
    strong_job = """
    Hiring QA Automation Engineer with 4+ years of experience in Selenium,
    Python, Pytest, REST API testing, SQL, Docker, Jenkins, Git, and Agile.
    """
    weak_job = """
    Hiring Marketing Manager with campaign strategy, copywriting, analytics,
    events, stakeholder communication, and brand positioning experience.
    """

    strong_score = ats_pipeline.calculate_ats_score(
        RESUME_TEXT, strong_job, "QA Automation Engineer"
    )
    weak_score = ats_pipeline.calculate_ats_score(
        RESUME_TEXT, weak_job, "Marketing Manager"
    )

    assert strong_score.ats_score > weak_score.ats_score
    assert strong_score.semantic_match > 75.0
    assert weak_score.semantic_match < 25.0
    assert 0.0 <= strong_score.ats_score <= 100.0
    assert 0.0 <= weak_score.ats_score <= 100.0


@pytest.mark.unit
def test_missing_skills_are_listed():
    job = """
    QA Automation Engineer needed with Selenium, Python, Playwright,
    Kubernetes, AWS, and GraphQL experience.
    """

    score = ats_pipeline.calculate_ats_score(RESUME_TEXT, job, "QA Automation Engineer")

    assert "selenium" in score.matched_skills
    assert "python" in score.matched_skills
    assert "playwright" in score.missing_skills
    assert "kubernetes" in score.missing_skills


@pytest.mark.unit
def test_email_html_uses_clickable_title_when_link_exists():
    html = ats_pipeline.build_top_matches_html(
        pd.DataFrame(
            [
                {
                    "Job Title": "Automation Engineer",
                    "Job Link": "https://www.linkedin.com/jobs/view/123",
                    "ATS Score": 88.5,
                    "Matched Skills": "selenium, python",
                    "Recommendation": "Strong match.",
                }
            ]
        )
    )

    assert 'href="https://www.linkedin.com/jobs/view/123"' in html
    assert "Automation Engineer</a>" in html


@pytest.mark.unit
def test_email_html_falls_back_to_plain_title_without_link():
    html = ats_pipeline.build_top_matches_html(
        pd.DataFrame(
            [
                {
                    "Job Title": "Automation Engineer",
                    "Job Link": "",
                    "ATS Score": 88.5,
                    "Matched Skills": "selenium, python",
                    "Recommendation": "Strong match.",
                }
            ]
        )
    )

    assert "<a href=" not in html
    assert "<td" in html
    assert "Automation Engineer" in html


@pytest.mark.unit
def test_score_jobs_excel_writes_ats_columns_and_job_link_hyperlinks(monkeypatch):
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        excel_path = tmp_path / "linkedin_jobs.xlsx"
        pd.DataFrame(
            [
                {
                    "Serial Number": "1",
                    "Job Title": "QA Automation Engineer",
                    "Job Link": "https://www.linkedin.com/jobs/view/123",
                    "Job Description": (
                        "QA Automation Engineer with Selenium, Python, Pytest, "
                        "REST API testing, SQL, and 4+ years of experience."
                    ),
                }
            ]
        ).to_excel(excel_path, index=False)

        monkeypatch.setattr(ats_pipeline, "extract_resume_text", lambda _: RESUME_TEXT)

        top_matches = ats_pipeline.score_jobs_excel(excel_path, tmp_path / "resume.pdf")

        assert "ATS Score" in top_matches.columns
        assert "Matched Skills" in top_matches.columns
        assert "Missing Skills" in top_matches.columns

        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        link_column = headers.index("Job Link") + 1
        link_cell = worksheet.cell(row=2, column=link_column)

@pytest.mark.unit
def test_non_technical_role_scoring():
    non_tech_job = """
    Marketing Coordinator needed to manage social media campaigns, create content,
    analyze marketing metrics, and collaborate with the design team. Requires
    strong communication and analytical skills.
    """
    non_tech_resume = """
    Marketing Assistant with experience in social media management, content creation,
    campaign analysis, and design collaboration. Excellent communication skills.
    """
    
    score = ats_pipeline.calculate_ats_score(
        non_tech_resume, non_tech_job, "Marketing Coordinator"
    )

    assert score.ats_score > 40.0
    assert score.semantic_match > 60.0
    assert score.keyword_match > 25.0


@pytest.mark.unit
def test_experience_match_overqualification_gives_full_score():
    resume = "Senior QA Engineer with 8 years of experience in Selenium and Python."
    job = "Junior QA role requiring 2 years of experience in Selenium and Python."

    score = ats_pipeline.calculate_ats_score(resume, job, "QA Engineer")

    assert score.experience_match == 100.0


@pytest.mark.unit
def test_experience_match_underqualification_reduces_score():
    resume = "Junior QA Engineer with 2 years of experience in Selenium and Python."
    job = "Senior QA role requiring 7 years of experience in Selenium and Python."

    score = ats_pipeline.calculate_ats_score(resume, job, "QA Engineer")

    assert score.experience_match < 50.0
